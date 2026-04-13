import os   
import json
import sys
import threading
import pandas
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime, timedelta

# 判断是否为打包环境
if getattr(sys, 'frozen', False):
    # 打包后的资源路径 (sys._MEIPASS)
    template_folder = os.path.join(sys._MEIPASS, 'templates')
    static_folder = os.path.join(sys._MEIPASS, 'static')
    app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
    # 输出目录在 exe 同级
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # 开发环境
    app = Flask(__name__)
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 配置输出目录
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')
# 定义子目录结构
DIRS = {
    'root': OUTPUT_DIR,
    'scores': os.path.join(OUTPUT_DIR, 'scores'),
    'teams': os.path.join(OUTPUT_DIR, 'teams'),
    'info': os.path.join(OUTPUT_DIR, 'match_info')
}

class Create_Scoretable:
    template_path = os.path.join(BASE_DIR, 'templates', 'scoretable_template.xlsx')

    def __init__(self, match_log_path=None):
        self.match_log_path = match_log_path if match_log_path is not None else []
        self.get_match_data() # 获取比赛数据

        self.wb = load_workbook(self.template_path) # 读取模板
        self.ws = self.wb.active # 选择活动表

        self.add_metadata() # 添加元数据
        self.add_scores() # 添加比分数据

        self.output_path = os.path.join(BASE_DIR, '记分表_' + (self.eventName or '') + datetime.now().strftime("%Y%m%d%H%M%S") + '.xlsx')
        self.wb.save(self.output_path) # 保存输出文件

    def get_match_data(self):
        try:
            with open(self.match_log_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            self.match_log = data
        except FileNotFoundError:
            print(f"文件不存在: {self.match_log_path}")
            return []
        except json.JSONDecodeError as e:
            print(f"JSON解析错误: {e}")
            return []
        except Exception as e:
            print(f"读取文件时出错: {e}")
            return []
        
        # 提取元数据
        self.metadata = self.match_log[0]["details"]
        # 提取赛事信息
        self.endTime = datetime.strptime(self.metadata["match_info"].get("endTime", "") or self.match_log[-1].get("timestamp", ""), "%Y-%m-%dT%H:%M")
        self.eventName = self.metadata["match_info"].get("eventName", "") or "N/A"
        self.serviceJudge = self.metadata["match_info"].get("serviceJudge", "") or "N/A"
        self.stage = self.metadata["match_info"].get("stage", "") or "N/A"
        self.startTime = datetime.strptime(self.metadata["match_info"].get("startTime", "") or "2026-01-01T00:00", "%Y-%m-%dT%H:%M")
        self.match_duation = self.endTime - self.startTime
        self.umpire = self.metadata["match_info"].get("umpire", "") or "N/A"
        self.venue = self.metadata["match_info"].get("venue", "") or "N/A"
        # 提取选手信息
        self.playerA1 = self.metadata["team_a"].get("p1", "")
        self.playerA2 = self.metadata["team_a"].get("p2", "")
        self.teamA = self.metadata["team_a"].get("name", "")
        self.playerB1 = self.metadata["team_b"].get("p1", "")
        self.playerB2 = self.metadata["team_b"].get("p2", "")
        self.teamB = self.metadata["team_b"].get("name", "")

        # 提取每局最终比分
        self.final_scores = [entry for entry in self.match_log if entry.get("type") == "set_end"]
        self.final_scores_str = [f"{_['details']['scoreA']}-{_['details']['scoreB']}" for _ in self.final_scores]

        # 处理比分日志，考虑撤销操作
        # 只保留 point 和 undo 操作，以应对局中可能的其他手动操作干扰
        process_log = [entry for entry in self.match_log if entry.get("type") in ["point", "undo"]]

        valid_events = []
        for entry in process_log:
            if entry.get("type") == "undo":
                if valid_events:
                    valid_events.pop()
            else:
                valid_events.append(entry)

        self.scores = valid_events

    def new_method(self):
        pass

    def add_metadata(self):
        self.ws['F4'] = self.eventName
        self.ws['F6'] = self.venue
        self.ws['F7'] = f"{self.startTime.month}.{self.startTime.day} {self.startTime.hour:02}:{self.startTime.minute:02}"
        self.ws['AQ4'] = self.umpire
        self.ws['AR5'] = self.serviceJudge
        self.ws['AP6'] = f"{self.startTime.hour:02}:{self.startTime.minute:02}"
        self.ws['AT6'] = f"{self.endTime.hour:02}:{self.endTime.minute:02}"
        self.ws['AR7'] = int(self.match_duation.total_seconds() / 60 + 0.5)

        self.ws['M5'] = self.playerA1
        self.ws['M6'] = self.playerA2
        self.ws['M7'] = self.teamA
        self.ws['AB5'] = self.playerB1
        self.ws['AB6'] = self.playerB2
        self.ws['AB7'] = self.teamB

        # 填充选手名单到每局
        for i in range(5):
            self.ws[f'B{9+i*5}'] = self.playerA1
            self.ws[f'B{10+i*5}'] = self.playerA2
            self.ws[f'B{11+i*5}'] = self.playerB1
            self.ws[f'B{12+i*5}'] = self.playerB2

    def get_player_row_offset(self, team, player_name):
        if team == "A":
            if player_name == self.playerA1: return 0
            if player_name == self.playerA2: return 1
        elif team == "B":
            if player_name == self.playerB1: return 2
            if player_name == self.playerB2: return 3
        return 0

    def get_write_pos(self, row, col):
        # 检查是否为合并单元格（需要跳过）
        while isinstance(self.ws.cell(row=row, column=col), MergedCell):
                col += 1
        
        # 计算当前单元格跨度
        span = 1
        for rng in self.ws.merged_cells.ranges:
            if row == rng.min_row and col == rng.min_col:
                    span = rng.max_col - rng.min_col + 1
                    break
        return col, span

    def add_scores(self):
        # 写入最终比分
        for i, final_score in enumerate(self.final_scores_str):
            self.ws.cell(row=5+i, column=24, value=final_score)

        # 以下是详细比分
        col_indices = [3] * 5
        initialized_sets = set()

        for i, point in enumerate(self.scores):
            details = point['details']
            setsA = details.get('setsA', 0)
            setsB = details.get('setsB', 0)
            set_idx = setsA + setsB
            
            if set_idx >= 5: continue

            row_base = 9 + set_idx * 5

            # 如果该局还没初始化（写入0-0），则先写入
            if set_idx not in initialized_sets:
                # 获取该局开始时的发球方和接发方（即当前分数的 server/receiver）
                # 注意：这里的 Server/Receiver 是指 0-0 这一球的
                start_server_team = details['serverTeam']
                start_server_player = details['serverPlayer']
                start_receiver_team = details['receiverTeam']
                start_receiver_player = details['receiverPlayer']

                s_row_offset = self.get_player_row_offset(start_server_team, start_server_player)
                r_row_offset = self.get_player_row_offset(start_receiver_team, start_receiver_player)
                
                s_row = row_base + s_row_offset
                r_row = row_base + r_row_offset

                # --- 步骤 1: 写入 S 和 R ---
                current_col = col_indices[set_idx]
                
                # Server 写 S
                s_col, s_span = self.get_write_pos(s_row, current_col)
                self.ws.cell(row=s_row, column=s_col, value="S")

                # Receiver 写 R
                r_col, r_span = self.get_write_pos(r_row, current_col)
                self.ws.cell(row=r_row, column=r_col, value="R")
                
                # 更新列索引 (跳过 S/R 所在的列)
                # 取最大跨度以确保两个单元格都被覆盖
                # (通常 s_col + s_span 应该等于 r_col + r_span，如果表格是对齐的)
                next_col_start = max(s_col + s_span, r_col + r_span)
                col_indices[set_idx] = next_col_start

                # --- 步骤 2: 写入 0 和 0 ---
                current_col = col_indices[set_idx]

                # Server 写 0
                s_col, s_span = self.get_write_pos(s_row, current_col)
                self.ws.cell(row=s_row, column=s_col, value="0")

                # Receiver 写 0
                r_col, r_span = self.get_write_pos(r_row, current_col)
                self.ws.cell(row=r_row, column=r_col, value="0")
                
                # 更新列索引 (跳过 0/0 所在的列)
                col_indices[set_idx] = max(s_col + s_span, r_col + r_span)
                
                initialized_sets.add(set_idx)

            winner = details['winner']
            score_val = details['newScoreA'] if winner == 'A' else details['newScoreB']
            
            target_row_offset = 0
            
            server_team = details['serverTeam']
            server_player = details['serverPlayer']

            # 如果是发球方得分，直接记在当前发球者行
            if winner == server_team:
                target_row_offset = self.get_player_row_offset(server_team, server_player)
            else:
                # 接发球方得分（换发球），则查看下一球的发球者
                next_point = None
                if i + 1 < len(self.scores):
                    next_p = self.scores[i+1]
                    n_setsA = next_p['details'].get('setsA', 0)
                    n_setsB = next_p['details'].get('setsB', 0)
                    if (n_setsA + n_setsB) == set_idx:
                         next_point = next_p

                if next_point:
                    next_server_team = next_point['details']['serverTeam']
                    next_server_player = next_point['details']['serverPlayer']
                    target_row_offset = self.get_player_row_offset(next_server_team, next_server_player)
                else:
                    # 最后一球且是接发方得分，记在接发者行
                    receiver_team = details['receiverTeam']
                    receiver_player = details['receiverPlayer']
                    target_row_offset = self.get_player_row_offset(receiver_team, receiver_player)
            
            row = row_base + target_row_offset
            base_col = col_indices[set_idx]
            
            # 获取写入位置和跨度
            col, span = self.get_write_pos(row, base_col)

            self.ws.cell(row=row, column=col, value=score_val)
            
            # 更新该局的列索引
            col_indices[set_idx] = col + span

# Create_Scoretable(r"D:\软件\小工具\Python_demo\比赛计分系统\FlyScore_Badminton_Live\match_log.json")
Create_Scoretable(input("请输入比赛记录文件路径（回车使用默认路径）：") or r"D:\软件\小工具\Python_demo\比赛计分系统\FlyScore_Badminton_Live\match_log.json")