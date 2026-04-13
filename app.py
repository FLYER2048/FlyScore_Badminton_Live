from flask import Flask, render_template, request, jsonify, send_file
import os
import json
import sys
import threading

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from datetime import datetime

# 判断是否为打包环境
if getattr(sys, 'frozen', False):
    # 打包后的资源路径 (sys._MEIPASS)
    template_folder = os.path.join(sys._MEIPASS, 'templates')
    static_folder = os.path.join(sys._MEIPASS, 'static')
    app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
    # 输出目录在 exe 同级
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # 开发环境
    app = Flask(__name__)
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 配置输出目录
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')

# 定义子目录结构
DIRS = {
    'root': OUTPUT_DIR,
    'scores': os.path.join(OUTPUT_DIR, 'scores'),
    'teams': os.path.join(OUTPUT_DIR, 'teams'),
    'info': os.path.join(OUTPUT_DIR, 'match_info')
}

# 确保所有目录存在
for d in DIRS.values():
    if not os.path.exists(d):
        os.makedirs(d)

GAME_STATE_FILE = os.path.join(OUTPUT_DIR, 'game_state.json')
MATCH_LOG_FILE = os.path.join(OUTPUT_DIR, 'match_log.json')
log_lock = threading.Lock()

class CreateScoretable:
    template_path = os.path.join(BASE_DIR, 'templates', 'scoretable_template.xlsx')
    DEFAULT_DATETIME = "2026-01-01T00:00"

    def __init__(self, match_log_path=None):
        self.match_log_path = match_log_path if match_log_path is not None else MATCH_LOG_FILE
        self.get_match_data() # 获取比赛数据

        self.wb = load_workbook(self.template_path) # 读取模板
        self.ws = self.wb.active # 选择活动表

        self.add_metadata() # 添加元数据
        self.add_scores() # 添加比分数据

        self.output_path = os.path.join(BASE_DIR, '记分表_' + (self.eventName.replace('/', '_') or 'NA') + "_" + datetime.now().strftime("%Y%m%d%H%M%S") + '.xlsx')
        self.wb.save(self.output_path) # 保存输出文件

    def get_match_data(self):
        try:
            with open(self.match_log_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 确保数据格式正确且非空
            if not isinstance(data, list) or not data:
                raise ValueError(f"无效的比赛数据: {self.match_log_path}")
            
            self.match_log = data
        except FileNotFoundError:
            print(f"文件不存在: {self.match_log_path}")
            # 重新抛出异常，防止后续逻辑在无效状态下继续执行
            raise
        except json.JSONDecodeError as e:
            print(f"JSON解析错误: {e}")
            # 重新抛出异常，防止后续逻辑在无效状态下继续执行
            raise
        except Exception as e:
            print(f"读取文件时出错: {e}")
            # 重新抛出异常，防止后续逻辑在无效状态下继续执行
            raise
        
        # 提取元数据
        self.metadata = self.match_log[0]["details"]
        # 提取赛事信息
        end_time_str = self.metadata["match_info"].get("endTime", "") or self.match_log[-1].get("timestamp", "")
        if end_time_str:
            try:
                self.endTime = datetime.strptime(end_time_str, "%Y-%m-%dT%H:%M")
            except ValueError:
                # 如果时间格式不符合预期，则使用默认时间避免程序崩溃
                self.endTime = datetime.strptime(self.DEFAULT_DATETIME, "%Y-%m-%dT%H:%M")
        else:
            # 当 endTime 和 timestamp 都为空时，使用默认时间
            self.endTime = datetime.strptime(self.DEFAULT_DATETIME, "%Y-%m-%dT%H:%M")
        
        self.eventName = self.metadata["match_info"].get("eventName", "") or "N/A"
        self.serviceJudge = self.metadata["match_info"].get("serviceJudge", "") or "N/A"
        self.match_type = self.metadata["match_info"].get("matchType", "") or "N/A"
        
        start_time_str = self.metadata["match_info"].get("startTime", "") or self.DEFAULT_DATETIME
        try:
            self.startTime = datetime.strptime(start_time_str, "%Y-%m-%dT%H:%M")
        except ValueError:
            # 如果时间格式不符合预期，则使用默认时间避免程序崩溃
            self.startTime = datetime.strptime(self.DEFAULT_DATETIME, "%Y-%m-%dT%H:%M")
        
        self.match_duration = self.endTime - self.startTime
        self.umpire = self.metadata["match_info"].get("umpire", "") or "N/A"
        self.venue = self.metadata["match_info"].get("venue", "") or "N/A"
        # 提取选手信息
        self.playerA1 = self.metadata["team_a"].get("p1", "")
        self.playerA2 = self.metadata["team_a"].get("p2", "")
        self.teamA = self.metadata["team_a"].get("name", "")
        self.playerB1 = self.metadata["team_b"].get("p1", "")
        self.playerB2 = self.metadata["team_b"].get("p2", "")
        self.teamB = self.metadata["team_b"].get("name", "")

        # 提取每局最终比分
        self.final_scores = [entry for entry in self.match_log if entry.get("type") == "set_end"]
        self.final_scores_str = [f"{_['details']['scoreA']}-{_['details']['scoreB']}" for _ in self.final_scores]

        # 处理比分日志，考虑撤销操作
        # 只保留 point 和 undo 操作，以应对局中可能的其他手动操作干扰
        process_log = [entry for entry in self.match_log if entry.get("type") in ["point", "undo"]]

        valid_events = []
        for entry in process_log:
            if entry.get("type") == "undo":
                if valid_events:
                    valid_events.pop()
            else:
                valid_events.append(entry)

        self.scores = valid_events

    def add_metadata(self):
        self.ws['F4'] = self.eventName
        self.ws['F5'] = self.match_type
        self.ws['F6'] = self.venue
        self.ws['F7'] = f"{self.startTime.month}.{self.startTime.day} {self.startTime.hour:02}:{self.startTime.minute:02}"
        self.ws['AQ4'] = self.umpire
        self.ws['AR5'] = self.serviceJudge
        self.ws['AP6'] = f"{self.startTime.hour:02}:{self.startTime.minute:02}"
        self.ws['AT6'] = f"{self.endTime.hour:02}:{self.endTime.minute:02}"
        self.ws['AR7'] = int(self.match_duration.total_seconds() / 60 + 0.5)

        self.ws['M5'] = self.playerA1
        self.ws['M6'] = self.playerA2
        self.ws['M7'] = self.teamA
        self.ws['AB5'] = self.playerB1
        self.ws['AB6'] = self.playerB2
        self.ws['AB7'] = self.teamB

        # 填充选手名单到每局
        for i in range(5):
            self.ws[f'B{9+i*5}'] = self.playerA1
            self.ws[f'B{10+i*5}'] = self.playerA2
            self.ws[f'B{11+i*5}'] = self.playerB1
            self.ws[f'B{12+i*5}'] = self.playerB2

    def get_player_row_offset(self, team, player_name):
        if team == "A":
            if player_name == self.playerA1: return 0
            if player_name == self.playerA2: return 1
        elif team == "B":
            if player_name == self.playerB1: return 2
            if player_name == self.playerB2: return 3
        return 0

    def get_write_pos(self, row, col):
        # 检查是否为合并单元格（需要跳过）
        while isinstance(self.ws.cell(row=row, column=col), MergedCell):
            col += 1
        
        # 计算当前单元格跨度
        span = 1
        for rng in self.ws.merged_cells.ranges:
            if row == rng.min_row and col == rng.min_col:
                span = rng.max_col - rng.min_col + 1
                break
        return col, span

    def add_scores(self):
        # 写入最终比分
        for i, final_score in enumerate(self.final_scores_str):
            self.ws.cell(row=5+i, column=24, value=final_score)

        # 以下是详细比分
        col_indices = [3] * 5
        initialized_sets = set()

        for i, point in enumerate(self.scores):
            details = point['details']
            setsA = details.get('setsA', 0)
            setsB = details.get('setsB', 0)
            set_idx = setsA + setsB
            
            if set_idx >= 5: continue

            row_base = 9 + set_idx * 5

            # 如果该局还没初始化（写入0-0），则先写入
            if set_idx not in initialized_sets:
                # 获取该局开始时的发球方和接发方（即当前分数的 server/receiver）
                # 注意：这里的 Server/Receiver 是指 0-0 这一球的
                start_server_team = details['serverTeam']
                start_server_player = details['serverPlayer']
                start_receiver_team = details['receiverTeam']
                start_receiver_player = details['receiverPlayer']

                s_row_offset = self.get_player_row_offset(start_server_team, start_server_player)
                r_row_offset = self.get_player_row_offset(start_receiver_team, start_receiver_player)
                
                s_row = row_base + s_row_offset
                r_row = row_base + r_row_offset

                # --- 步骤 1: 写入 S 和 R ---
                current_col = col_indices[set_idx]
                
                # Server 写 S
                s_col, s_span = self.get_write_pos(s_row, current_col)
                self.ws.cell(row=s_row, column=s_col, value="S")

                # Receiver 写 R
                r_col, r_span = self.get_write_pos(r_row, current_col)
                self.ws.cell(row=r_row, column=r_col, value="R")
                
                # 更新列索引 (跳过 S/R 所在的列)
                # 取最大跨度以确保两个单元格都被覆盖
                # (通常 s_col + s_span 应该等于 r_col + r_span，如果表格是对齐的)
                next_col_start = max(s_col + s_span, r_col + r_span)
                col_indices[set_idx] = next_col_start

                # --- 步骤 2: 写入 0 和 0 ---
                current_col = col_indices[set_idx]

                # Server 写 0
                s_col, s_span = self.get_write_pos(s_row, current_col)
                self.ws.cell(row=s_row, column=s_col, value="0")

                # Receiver 写 0
                r_col, r_span = self.get_write_pos(r_row, current_col)
                self.ws.cell(row=r_row, column=r_col, value="0")
                
                # 更新列索引 (跳过 0/0 所在的列)
                col_indices[set_idx] = max(s_col + s_span, r_col + r_span)
                
                initialized_sets.add(set_idx)

            winner = details['winner']
            score_val = details['newScoreA'] if winner == 'A' else details['newScoreB']
            
            target_row_offset = 0
            
            server_team = details['serverTeam']
            server_player = details['serverPlayer']

            # 如果是发球方得分，直接记在当前发球者行
            if winner == server_team:
                target_row_offset = self.get_player_row_offset(server_team, server_player)
            else:
                # 接发球方得分（换发球），则查看下一球的发球者
                next_point = None
                if i + 1 < len(self.scores):
                    next_p = self.scores[i+1]
                    n_setsA = next_p['details'].get('setsA', 0)
                    n_setsB = next_p['details'].get('setsB', 0)
                    if (n_setsA + n_setsB) == set_idx:
                         next_point = next_p

                if next_point:
                    next_server_team = next_point['details']['serverTeam']
                    next_server_player = next_point['details']['serverPlayer']
                    target_row_offset = self.get_player_row_offset(next_server_team, next_server_player)
                else:
                    # 最后一球且是接发方得分，记在接发者行
                    receiver_team = details['receiverTeam']
                    receiver_player = details['receiverPlayer']
                    target_row_offset = self.get_player_row_offset(receiver_team, receiver_player)
            
            row = row_base + target_row_offset
            base_col = col_indices[set_idx]
            
            # 获取写入位置和跨度
            col, span = self.get_write_pos(row, base_col)

            self.ws.cell(row=row, column=col, value=score_val)
            
            # 更新该局的列索引
            col_indices[set_idx] = col + span

def write_txt(category, filename, content):
    """
    category: 'scores', 'teams', 'info'
    """
    target_dir = DIRS.get(category, OUTPUT_DIR)
    try:
        with open(os.path.join(target_dir, filename), 'w', encoding='utf-8') as f:
            f.write(str(content))
    except Exception as e:
        print(f"Error writing to {category}/{filename}: {e}")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/scoreboard')
def scoreboard():
    return render_template('scoreboard.html')

@app.route('/api/get_state', methods=['GET'])
def get_state():
    if os.path.exists(GAME_STATE_FILE):
        try:
            with open(GAME_STATE_FILE, 'r', encoding='utf-8') as f:
                return jsonify(json.load(f))
        except Exception as e:
            print(f"Error reading state: {e}")
            return jsonify({})
    return jsonify({})

@app.route('/api/update_status', methods=['POST'])
def update_status():
    data = request.json
    
    # 保存完整状态用于恢复
    if 'fullState' in data:
        try:
            with open(GAME_STATE_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error saving state: {e}")

    # 写入文件
    
    # 1. 赛事元数据 (match_info)
    write_txt('info', 'event_name.txt', data.get('event_name', ''))
    write_txt('info', 'match_stage.txt', data.get('match_stage', ''))
    write_txt('info', 'match_venue.txt', data.get('match_venue', ''))
    write_txt('info', 'umpire.txt', data.get('umpire', ''))
    write_txt('info', 'service_judge.txt', data.get('service_judge', ''))
    write_txt('info', 'start_time.txt', data.get('start_time', ''))
    write_txt('info', 'end_time.txt', data.get('end_time', ''))
    write_txt('info', 'match_status.txt', data.get('status_message', ''))

    # 2. 比赛双方 (teams)
    write_txt('teams', 'team_a_name.txt', data.get('team_a_name', ''))
    write_txt('teams', 'team_b_name.txt', data.get('team_b_name', ''))
    
    # 发球方指示
    if data.get('serving_team') == 'A':
        write_txt('teams', 'indicator_a.txt', '🏸')
        write_txt('teams', 'indicator_b.txt', '')
    else:
        write_txt('teams', 'indicator_a.txt', '')
        write_txt('teams', 'indicator_b.txt', '🏸')
    
    # 3. 比分 (scores)
    # 大比分 (局分)
    write_txt('scores', 'score_a_sets.txt', data.get('sets_a', 0))
    write_txt('scores', 'score_b_sets.txt', data.get('sets_b', 0))
    write_txt('scores', 'score_sets_combined.txt', f"{data.get('sets_a', 0)} - {data.get('sets_b', 0)}")
    
    # 小比分 (当前局分数)
    write_txt('scores', 'score_a_points.txt', data.get('points_a', 0))
    write_txt('scores', 'score_b_points.txt', data.get('points_b', 0))
    write_txt('scores', 'score_points_combined.txt', f"{data.get('points_a', 0)} - {data.get('points_b', 0)}")

    return jsonify({"status": "success"})

@app.route('/api/log_event', methods=['POST'])
def log_event():
    event = request.json
    
    with log_lock:
        # 如果是比赛开始事件，清空日志
        if event.get('type') == 'match_start':
            logs = [event]
            try:
                with open(MATCH_LOG_FILE, 'w', encoding='utf-8') as f:
                    json.dump(logs, f, ensure_ascii=False, indent=2)
                    f.flush()
                    os.fsync(f.fileno())
                return jsonify({"status": "success", "message": "Log reset"})
            except Exception as e:
                print(f"Error resetting log: {e}")
                return jsonify({"status": "error", "message": str(e)}), 500

        # 读取现有日志
        logs = []
        if os.path.exists(MATCH_LOG_FILE):
            try:
                with open(MATCH_LOG_FILE, 'r', encoding='utf-8') as f:
                    content = f.read()
                    # 移除 NUL 字符，防止文件损坏导致读取失败
                    content = content.replace('\x00', '')
                    if content.strip():
                        try:
                            logs = json.loads(content)
                        except json.JSONDecodeError:
                            print("JSON Decode Error in match_log.json, starting fresh or appending.")
                            # 如果解析失败，可能是文件截断。
                            # 这里为了简单起见，如果无法解析，我们尝试保留旧内容（如果需要更复杂的恢复逻辑可以加）
                            # 但为了保证程序不崩，我们初始化为空列表，这会导致旧日志丢失。
                            # 更好的做法可能是备份坏文件。
                            pass
            except Exception as e:
                print(f"Error reading log: {e}")
        
        logs.append(event)
        
        try:
            with open(MATCH_LOG_FILE, 'w', encoding='utf-8') as f:
                json.dump(logs, f, ensure_ascii=False, indent=2)
                f.flush()
                os.fsync(f.fileno())
        except Exception as e:
            print(f"Error writing log: {e}")
            return jsonify({"status": "error", "message": str(e)}), 500
            
        return jsonify({"status": "success"})

@app.route('/api/download_log')
def download_log():
    if os.path.exists(MATCH_LOG_FILE):
        return send_file(MATCH_LOG_FILE, as_attachment=True, download_name='match_log.json')
    else:
        return jsonify({"error": "Log file not found"}), 404

@app.route('/api/generate_scoretable', methods=['POST'])
def generate_scoretable():
    try:
        # 使用当前的 MATCH_LOG_FILE
        generator = CreateScoretable()
        # 获取生成的文件名
        filename = os.path.basename(generator.output_path)
        return jsonify({"status": "success", "filename": filename})
    except Exception as e:
        print(f"Error generating scoretable: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/download_scoretable/<filename>')
def download_scoretable(filename):
    # 确保文件名安全：将输入限制为基础文件名并拒绝路径遍历
    safe_filename = os.path.basename(filename)
    # 如果归一化后文件名发生变化或为空，则视为非法
    if not safe_filename or safe_filename != filename or safe_filename in (".", ".."):
        return jsonify({"error": "Invalid filename"}), 400
    file_path = os.path.join(BASE_DIR, safe_filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, download_name=safe_filename)
    else:
        return jsonify({"error": "File not found"}), 404

if __name__ == '__main__':
    # 打包后禁用 debug
    debug_mode = not getattr(sys, 'frozen', False)
    if not debug_mode:
        print("FlyScore Server is running on http://127.0.0.1:5000")
        print("Close this window to stop the server.")
    app.run(debug=debug_mode, host='0.0.0.0', port=5000)
